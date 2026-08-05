#!/usr/bin/env python3

import sqlite3
import sys
from pathlib import Path
from openpyxl import load_workbook

DATABASE = "rates.db"

# Default Excel file
if len(sys.argv) > 1:
    excel_file = Path(sys.argv[1])
else:
    excel_file = Path("data/Effective_Towing_Rates.xlsx")

if not excel_file.exists():
    print(f"Error: {excel_file} not found.")
    sys.exit(1)

print(f"Reading {excel_file}...")

wb = load_workbook(excel_file, data_only=True)
ws = wb.active

rows = list(ws.iter_rows(values_only=True))

headers = list(rows[0])

headers.insert(0, "ID")

data = []

counter = 1

for row in rows[1:]:

    row = list(row)

    while len(row) < len(headers) - 1:
        row.append(None)

    row.insert(0, counter)

    data.append(row)

    counter += 1

conn = sqlite3.connect(DATABASE)

cur = conn.cursor()

cur.execute("DROP TABLE IF EXISTS rates")

columns = []

for header in headers:
    if header == "ID":
        columns.append('"ID" INTEGER PRIMARY KEY')
    else:
        columns.append(f'"{header}" TEXT')

create_sql = f"""
CREATE TABLE rates (
{', '.join(columns)}
)
"""

cur.execute(create_sql)

placeholders = ",".join(["?"] * len(headers))

insert_sql = f"""
INSERT INTO rates
VALUES ({placeholders})
"""

cur.executemany(insert_sql, data)

conn.commit()

conn.close()

print(f"Imported {len(data)} records.")
print("Database created successfully.")
